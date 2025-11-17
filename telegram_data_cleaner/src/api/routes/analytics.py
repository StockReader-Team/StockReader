"""
API routes for channel analytics.
"""
from datetime import datetime, timedelta, date
from typing import Optional, List
import uuid
import io

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.database import db_manager
from src.core.analytics.channel_analytics_service import ChannelAnalyticsService
from src.core.logging import get_logger

logger = get_logger(__name__)

router = APIRouter(prefix="/analytics", tags=["Analytics"])


async def get_db_session() -> AsyncSession:
    """Get database session dependency."""
    async with db_manager.session() as session:
        yield session


@router.get("/channels/{channel_id}/stats")
async def get_channel_stats(
    channel_id: str,
    time_range: str = Query(
        "30min",
        description="Time range: 5min, 30min, 1hour, today, 7days, 30days"
    ),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get channel statistics for a specific time range.

    Args:
        channel_id: Channel UUID
        time_range: Time range to analyze
        session: Database session

    Returns:
        Channel statistics including message counts, matches, and top items
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)

    # Parse time range
    if time_range == "5min":
        stats = await analytics_service.get_realtime_stats(channel_uuid, minutes=5)
    elif time_range == "30min":
        stats = await analytics_service.get_realtime_stats(channel_uuid, minutes=30)
    elif time_range == "1hour":
        stats = await analytics_service.get_realtime_stats(channel_uuid, minutes=60)
    elif time_range in ["today", "7days", "30days"]:
        # Get from aggregates
        days_map = {"today": 1, "7days": 7, "30days": 30}
        days = days_map[time_range]
        stats = await analytics_service.get_channel_content_profile(
            channel_uuid, days=days
        )
    else:
        raise HTTPException(status_code=400, detail="Invalid time_range parameter")

    return stats


@router.get("/channels/{channel_id}/content-profile")
async def get_content_profile(
    channel_id: str,
    days: int = Query(7, ge=1, le=90, description="Number of days to analyze"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get content profile for a channel showing what topics it focuses on.

    Args:
        channel_id: Channel UUID
        days: Number of days to analyze
        session: Database session

    Returns:
        Content profile with category distribution
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)
    profile = await analytics_service.get_channel_content_profile(
        channel_uuid, days=days
    )

    return profile


@router.get("/channels/compare")
async def compare_channels(
    channel_ids: str = Query(
        ...,
        description="Comma-separated list of channel UUIDs"
    ),
    days: int = Query(7, ge=1, le=90, description="Number of days to analyze"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Compare multiple channels.

    Args:
        channel_ids: Comma-separated channel UUIDs
        days: Number of days to analyze
        session: Database session

    Returns:
        Comparison data for all channels
    """
    # Parse channel IDs
    try:
        channel_uuids = [uuid.UUID(cid.strip()) for cid in channel_ids.split(",")]
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Invalid channel ID format"
        )

    if len(channel_uuids) > 10:
        raise HTTPException(
            status_code=400,
            detail="Maximum 10 channels can be compared at once"
        )

    analytics_service = ChannelAnalyticsService(session)

    # Get profile for each channel
    channels_data = []
    for channel_uuid in channel_uuids:
        profile = await analytics_service.get_channel_content_profile(
            channel_uuid, days=days
        )
        channels_data.append(profile)

    return {
        "days": days,
        "channels": channels_data
    }


@router.get("/channels/{channel_id}/export/excel")
async def export_channel_analytics_excel(
    channel_id: str,
    days: int = Query(7, ge=1, le=90, description="Number of days to export"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Export channel analytics to Excel file.

    Args:
        channel_id: Channel UUID
        days: Number of days to export
        session: Database session

    Returns:
        Excel file as streaming response
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)

    # Get data
    profile = await analytics_service.get_channel_content_profile(
        channel_uuid, days=days
    )
    realtime_stats = await analytics_service.get_realtime_stats(
        channel_uuid, minutes=60
    )

    # Create Excel workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Overview sheet
    ws_overview = wb.create_sheet("Overview")
    _create_overview_sheet(ws_overview, profile, realtime_stats)

    # Create Categories sheet
    ws_categories = wb.create_sheet("Categories")
    _create_categories_sheet(ws_categories, profile)

    # Create Symbols sheet
    ws_symbols = wb.create_sheet("Top Symbols")
    _create_top_items_sheet(ws_symbols, realtime_stats.get("top_symbols", []), "Symbols")

    # Create Industries sheet
    ws_industries = wb.create_sheet("Top Industries")
    _create_top_items_sheet(ws_industries, realtime_stats.get("top_industries", []), "Industries")

    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return as streaming response
    filename = f"channel_analytics_{channel_id}_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _create_overview_sheet(ws, profile, realtime_stats):
    """Create overview sheet in Excel."""
    # Title
    ws["A1"] = "Channel Analytics Overview"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:B1")

    # Add data
    row = 3
    data = [
        ("Analysis Period (days)", profile.get("days", 0)),
        ("Total Messages", profile.get("total_messages", 0)),
        ("Total Matches", profile.get("total_matches", 0)),
        ("Primary Focus", profile.get("primary_focus", "N/A")),
        ("Focus Percentage", f"{profile.get('focus_percentage', 0)}%"),
        ("", ""),
        ("Last Hour Stats", ""),
        ("Messages (last hour)", realtime_stats.get("message_count", 0)),
        ("Matches (last hour)", realtime_stats.get("match_count", 0)),
    ]

    for label, value in data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        if label and not value:
            ws[f"A{row}"].font = Font(bold=True)
        row += 1

    # Style
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _create_categories_sheet(ws, profile):
    """Create categories sheet in Excel."""
    # Headers
    headers = ["Category Name", "Match Count", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data
    categories = profile.get("categories", [])
    for row, category in enumerate(categories, start=2):
        ws.cell(row=row, column=1, value=category.get("name", ""))
        ws.cell(row=row, column=2, value=category.get("count", 0))
        ws.cell(row=row, column=3, value=f"{category.get('percentage', 0)}%")

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def _create_top_items_sheet(ws, items, item_type):
    """Create top items sheet (symbols/industries) in Excel."""
    # Headers
    if item_type == "Symbols":
        headers = ["Symbol", "Match Count"]
    else:
        headers = ["Industry", "Match Count"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row, item in enumerate(items, start=2):
        if item_type == "Symbols":
            ws.cell(row=row, column=1, value=item.get("word", ""))
        else:
            ws.cell(row=row, column=1, value=item.get("name", ""))
        ws.cell(row=row, column=2, value=item.get("count", 0))

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


@router.get("/overview")
async def get_global_overview(
    days: int = Query(7, ge=1, le=90, description="Number of days to analyze"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get global analytics overview across all channels.

    Args:
        days: Number of days to analyze
        session: Database session

    Returns:
        Global overview including:
        - Total stats (messages, matches, channels)
        - Most active channel
        - Top symbols, industries, and categories across all channels
        - List of all channels with their stats
    """
    analytics_service = ChannelAnalyticsService(session)
    overview = await analytics_service.get_global_overview(days=days)

    return overview


@router.post("/compute-aggregates")
async def compute_aggregates_now(
    hours_back: int = Query(1, ge=1, le=24, description="Hours to compute"),
    granularity: str = Query("hourly", description="hourly or daily"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Manually trigger analytics aggregation computation.

    Args:
        hours_back: Number of hours to compute backwards from now
        granularity: "hourly" or "daily"
        session: Database session

    Returns:
        Number of records created/updated
    """
    if granularity not in ["hourly", "daily"]:
        raise HTTPException(
            status_code=400,
            detail="granularity must be 'hourly' or 'daily'"
        )

    analytics_service = ChannelAnalyticsService(session)

    now = datetime.utcnow()
    start_time = now - timedelta(hours=hours_back)

    records_created = await analytics_service.compute_aggregates(
        start_datetime=start_time,
        end_datetime=now,
        granularity=granularity
    )

    return {
        "status": "success",
        "records_created": records_created,
        "start_time": start_time.isoformat(),
        "end_time": now.isoformat(),
        "granularity": granularity
    }


@router.get("/channels/{channel_id}/dictionary-words")
async def get_channel_dictionary_words(
    channel_id: str,
    dictionary_name: str = Query(..., description="Name of the dictionary"),
    days: int = Query(7, ge=1, le=90, description="Number of days to analyze"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get dictionary words used by a specific channel.

    Args:
        channel_id: Channel UUID
        dictionary_name: Name of the dictionary
        days: Number of days to analyze
        session: Database session

    Returns:
        List of words from the dictionary used in this channel with their counts
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)
    words = await analytics_service.get_channel_dictionary_words(
        channel_uuid, dictionary_name, days
    )

    return {"words": words}


@router.get("/channels/{channel_id}/timeline")
async def get_channel_timeline(
    channel_id: str,
    days: int = Query(15, ge=1, le=90, description="Number of days for timeline"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get timeline analytics for a channel showing daily stats over time.

    Args:
        channel_id: Channel UUID
        days: Number of days for timeline
        session: Database session

    Returns:
        Timeline data with daily breakdown of messages, matches, symbols, industries, categories
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)
    timeline = await analytics_service.get_channel_timeline(channel_uuid, days)

    return timeline


@router.get("/channels/{channel_id}/hourly-activity")
async def get_channel_hourly_activity(
    channel_id: str,
    days: int = Query(7, ge=1, le=30, description="Number of days to analyze"),
    session: AsyncSession = Depends(get_db_session)
):
    """
    Get hourly activity pattern for a channel.

    Args:
        channel_id: Channel UUID
        days: Number of days to analyze
        session: Database session

    Returns:
        Hourly activity data with message counts per hour and heatmap data
    """
    try:
        channel_uuid = uuid.UUID(channel_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid channel ID format")

    analytics_service = ChannelAnalyticsService(session)
    activity = await analytics_service.get_channel_hourly_activity(channel_uuid, days)

    return activity
